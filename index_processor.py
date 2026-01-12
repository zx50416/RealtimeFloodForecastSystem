import numpy as np
import openpyxl

def RMSE(obv, est):
    return np.sqrt(np.mean((est - obv) ** 2))

def MAE(obv, est):
    return np.mean(np.abs(est - obv))

def CE(obv, est):
    return 1 - np.sum((est - obv) ** 2) / np.sum((obv - np.mean(obv)) ** 2)

def CC(obv, est):
    return np.corrcoef(est, obv)[0, 1]

def EQp(obv, est):
    return (np.max(est) - np.max(obv)) / np.max(obv)

def ETp(obv, est):
    return np.abs(np.argmax(est) - np.argmax(obv))

def Caculate_Index(index_name, obv, est):
    if index_name == 'RMSE': return RMSE(obv, est)
    if index_name == 'MAE': return MAE(obv, est)
    if index_name == 'CE': return CE(obv, est)
    if index_name == 'CC': return CC(obv, est)
    if index_name == 'EQp': return EQp(obv, est)
    if index_name == 'ETp': return ETp(obv, est)

def get_all_indices(num_events, RES_train, RES_test, index_names):

    index_train = []
    index_test = []

    for i in range(num_events):
        obv_train = np.array(RES_train[i][0], dtype=np.float64)
        est_train = np.array(RES_train[i][1], dtype=np.float64)
        obv_test = np.array(RES_test[i][0], dtype=np.float64)
        est_test = np.array(RES_test[i][1], dtype=np.float64)

        for dtype in ('train', 'test'):
            temp = {}
            if dtype == 'train':
                for index_name in index_names:
                    temp[index_name] = Caculate_Index(index_name, obv_train, est_train)
                index_train.append(temp)
            elif dtype == 'test':
                for index_name in index_names:
                    temp[index_name] = Caculate_Index(index_name, obv_test, est_test)
                index_test.append(temp)

    return index_train, index_test

def write_Index(num_events, eventWithMaxVal, index_names, index_train, index_test, output_path):
    
    # 將指標寫入excel檔
    wb = openpyxl.Workbook()
    del wb['Sheet']

    # 計算表格數據(第一個工作表)
    avg_index_train = {}
    avg_index_test = {}

    for index_name in index_names:
        total_train = []
        total_test = []
        for i in range(num_events):
            if i == eventWithMaxVal-1: 
                continue
            total_train.append(index_train[i][index_name])
            total_test.append(index_test[i][index_name])
        avg_index_train[index_name] = np.mean(total_train)
        avg_index_test[index_name] = np.mean(total_test)
    
    # print('Avg index train:\n', avg_index_train)
    # print('Avg index test:\n', avg_index_test)

    sheet = wb.create_sheet("Event Statistics")
    sheet.append([ "Event with max value" ])
    sheet.append([ eventWithMaxVal ])
    sheet.append([ None, 'Train', 'Test' ])
    for index_name in index_names:
        sheet.append([ index_name, round(avg_index_train[index_name], 4), round(avg_index_test[index_name], 4) ])

    # 將不同指標寫入不同工作表
    for index_name in index_names:
        sheet = wb.create_sheet(index_name)
        sheet.append([ 'Event', 'Train', 'Test' ])
        for i in range(num_events):
            sheet.append([ i+1, round(index_train[i][index_name], 4), round(index_test[i][index_name], 4) ])

    # 輸出excel檔
    wb.save(output_path)
    print("Result saved in " + output_path)