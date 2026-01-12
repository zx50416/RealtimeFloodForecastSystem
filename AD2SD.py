import numpy as np
import pandas as pd
import openpyxl

IO = 0
FACTOR = 1
LAG = 2

# 取得資料值並儲存為列表
def get_values(sheet):
    arr = [] 
    for row in sheet:
        temp = []  
        for column in row:
            temp.append(column.value)
        arr.append(temp)
    return arr

# 讀入AD和IOLag
def read_AD(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    events = [wb[sheetname] for sheetname in wb.sheetnames]
    for i in range(len(events)):
        events[i] = pd.DataFrame(get_values(events[i])).dropna(axis=0) 
    return events

def read_IOLag(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    iolag = wb['IOLag']
    iolag = pd.DataFrame(get_values(iolag)).drop(0, axis=0)
    return iolag

# 轉為SD
def ad2sd(events, iolag):
    sd = []
    # 找出最大及最小lag
    maxShift = max(iolag[LAG])
    minShift = min(iolag[LAG])
    for event in events:
        tempEvent = []
        for i in range(len(iolag)):
            factor = iolag[FACTOR][i+1] - 1
            lag = iolag[LAG][i+1]
            temp = list(event[factor][(lag-minShift+1):(lag+len(event)-maxShift)])
            tempEvent.append(temp)
        tempEvent = np.array(tempEvent).T.tolist()
        sd.append(tempEvent)
    return sd

# 寫入excel
def write_SD(events, sd, path):
    wb = openpyxl.Workbook()
    del wb['Sheet']
    for i in range(len(events)):
        sheet = wb.create_sheet("Event " + str(i+1))
        for j in range(len(sd[i])):
            sheet.append(sd[i][j])
    wb.save(path) 
    print("File saved in " + path)

## main
# events = read_AD(adFile)
# iolag = read_IOLag(iolagFile)
# sd = ad2sd(events, iolag)
# write_SD(events, sd, sdFile)
