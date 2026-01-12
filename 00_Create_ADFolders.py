# 建立資料夾(同一空間點位)(T + delta T)，並建立IOLag.xlsx
import os
from openpyxl import Workbook
import shutil

PROJECT_FOLDER = 'C:/MyPython/01_InunFore/01-1_RegularModel_ParaOpt/'
base_path = os.path.join(PROJECT_FOLDER, 'DATA/')

for FutureTime in range(1, 7):
    # 创建文件夹
    folder_name = f'IP01-{FutureTime}'
    folder_path = os.path.join(base_path, folder_name)
    os.makedirs(folder_path, exist_ok=True)
    print(f'Created folder: {folder_path}')

    # 创建并写入Excel文件
    file_name = os.path.join(folder_path, 'IOLag.xlsx')
    wb = Workbook()
    ws = wb.active
    
    # 写入数据
    ws['A1'], ws['B1'], ws['C1'] = "IO", "Factor", "Lag(IOFL2)"
    ws['A2'], ws['B2'], ws['C2'] = 1, 25, 0
    ws['A3'], ws['B3'], ws['C3'] = 1, 25, -1
    ws['A4'], ws['B4'], ws['C4'] = 2, 25, FutureTime

    # 保存文件
    wb.save(file_name)
    print(f'Created file: {file_name}')







# ===========================================================
# 從A資料夾複製AD檔 至 其他建立的資料夾
import os
from openpyxl import Workbook
import shutil

# 決定空間點位
IPnumber = '01'

# A 資料夾路徑
folder_a = 'C:/MyPython/01_InunFore/01-1_RegularModel_ParaOpt/DATA/IP' + IPnumber + '-1'

# 目標資料夾的路徑
target_folders = [
    'C:/MyPython/01_InunFore/Test/DATA/WQP' + IPnumber + '-1',
    'C:/MyPython/01_InunFore/Test/DATA/WQP' + IPnumber + '-2',
    'C:/MyPython/01_InunFore/Test/DATA/WQP' + IPnumber + '-3',
    'C:/MyPython/Model_淹水預報/01_Regular Model/data/IP' + IPnumber + '-4',
    'C:/MyPython/Model_淹水預報/01_Regular Model/data/IP' + IPnumber + '-5',
    'C:/MyPython/Model_淹水預報/01_Regular Model/data/IP' + IPnumber + '-6'
]

# 檔案名稱
file_name = 'AD.xlsx'

# 複製檔案到目標資料夾
for target_folder in target_folders:
    target_path = os.path.join(target_folder, file_name)
    shutil.copy(f"{folder_a}/{file_name}", target_path)
    print(f"檔案 {file_name} 已複製到 {target_folder} 資料夾中")