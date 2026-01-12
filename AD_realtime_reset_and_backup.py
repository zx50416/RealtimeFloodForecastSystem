import os
import shutil
from datetime import datetime
from openpyxl import load_workbook


def backup_and_reset_ad_realtime():
    base_dir = "./_internal/DATA"
    src_file = os.path.join(base_dir, "AD_realtime.xlsx")
    history_dir = os.path.join(base_dir, "HISTORY_AD_realtime")

    # === 確保 HISTORY 資料夾存在 ===
    os.makedirs(history_dir, exist_ok=True)

    # === 備份原始檔案 ===
    now_str = datetime.now().strftime("%Y%m%d_%H%M")
    backup_name = f"AD_realtime_{now_str}.xlsx"
    backup_path = os.path.join(history_dir, backup_name)

    shutil.copy2(src_file, backup_path)

    # === 開啟原始 Excel（直接修改，不重建） ===
    wb = load_workbook(src_file)
    ws = wb.active  # 預設第一個工作表

    # === 清空第 2 列到最後一列（保留標題列）===
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None




    # === 存回原檔 ===
    wb.save(src_file)
    wb.close()

if __name__ == "__main__":
    backup_and_reset_ad_realtime()
    print("✅ AD_realtime.xlsx 已備份並重置完成")
    input("按 Enter 鍵結束程式...")