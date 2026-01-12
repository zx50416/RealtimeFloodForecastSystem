# merge_village_levels_to_final.py
# 功能：
#   讀 OUTPUTS/village_depth_for_risk.xlsx 的 H_1~H_4
#   用 (TOWNNAME, VILLNAME) 去對應 OUTPUTS/final_output.xlsx
#   只覆寫 H_1~H_4 四個欄位，其餘欄位與公式全部保留

import os
import pandas as pd
import openpyxl  # 要確定有安裝：pip install openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FOLDER = os.path.join(BASE_DIR, "OUTPUTS")

FINAL_FILE = os.path.join(OUTPUT_FOLDER, "final_output.xlsx")
VIL_DEPTH  = os.path.join(OUTPUT_FOLDER, "village_depth_for_risk.xlsx")


def merge_village_levels_to_final():
    if not os.path.exists(FINAL_FILE):
        raise FileNotFoundError("找不到 final_output.xlsx：" + FINAL_FILE)

    if not os.path.exists(VIL_DEPTH):
        raise FileNotFoundError("找不到 village_depth_for_risk.xlsx：" + VIL_DEPTH)

    # 1) 先用 pandas 讀 village_depth_for_risk.xlsx，拿 H_1~H_4
    df_vil = pd.read_excel(VIL_DEPTH)

    cols_needed = ["TOWNNAME", "VILLNAME", "H_1", "H_2", "H_3", "H_4"]
    for c in cols_needed:
        if c not in df_vil.columns:
            raise ValueError(f"village_depth_for_risk.xlsx 缺少欄位：{c}")

    # 建立一個 dict： (TOWNNAME, VILLNAME) -> (H1, H2, H3, H4)
    df_sub = df_vil[cols_needed].copy()
    mapping = {}
    for _, row in df_sub.iterrows():
        key = (str(row["TOWNNAME"]), str(row["VILLNAME"]))
        mapping[key] = (row["H_1"], row["H_2"], row["H_3"], row["H_4"])

    # 2) 用 openpyxl 開啟 final_output.xlsx，保留原始公式
    wb = openpyxl.load_workbook(FINAL_FILE, data_only=False)
    ws = wb.active  # 如果你的資料在別的工作表，就改成 wb["工作表名稱"]

    # 3) 找出標題列在哪些欄：TOWNNAME / VILLNAME / H_1~H_4
    header_row = 1
    col_index = {}

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val in ["TOWNNAME", "VILLNAME", "H_1", "H_2", "H_3", "H_4"]:
            col_index[str(val)] = col

    required_headers = ["TOWNNAME", "VILLNAME", "H_1", "H_2", "H_3", "H_4"]
    for h in required_headers:
        if h not in col_index:
            raise ValueError(f"final_output.xlsx 標題列找不到欄位：{h}")

    col_town = col_index["TOWNNAME"]
    col_vill = col_index["VILLNAME"]
    col_h1   = col_index["H_1"]
    col_h2   = col_index["H_2"]
    col_h3   = col_index["H_3"]
    col_h4   = col_index["H_4"]

    # 4) 逐列掃描 final_output.xlsx，依 (TOWNNAME, VILLNAME) 填值
    updated_rows = 0

    for row in range(header_row + 1, ws.max_row + 1):
        town = ws.cell(row=row, column=col_town).value
        vill = ws.cell(row=row, column=col_vill).value
        if town is None or vill is None:
            continue

        key = (str(town), str(vill))
        if key not in mapping:
            # 這個村里在 village_depth_for_risk 中沒有對應值，就跳過
            continue

        h1, h2, h3, h4 = mapping[key]

        ws.cell(row=row, column=col_h1, value=h1)
        ws.cell(row=row, column=col_h2, value=h2)
        ws.cell(row=row, column=col_h3, value=h3)
        ws.cell(row=row, column=col_h4, value=h4)

        updated_rows += 1

    wb.save(FINAL_FILE)
    print(f"✅ final_output.xlsx 已成功更新 H_1~H_4，共更新 {updated_rows} 列 → {FINAL_FILE}")
